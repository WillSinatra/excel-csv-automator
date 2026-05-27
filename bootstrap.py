"""
bootstrap.py — Excel CSV Automator Project Setup
Run once with: python bootstrap.py
Creates the full excel_csv_automator/ project in the current directory.
"""

from pathlib import Path
import textwrap

BASE = Path(__file__).parent / "excel_csv_automator"

FILES: dict[str, str] = {}

# ──────────────────────────────────────────────────────────────────────────────
# config.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["config.py"] = textwrap.dedent('''\
    """Global configuration for Excel CSV Automator."""

    from pathlib import Path

    # Directorio base del proyecto
    BASE_DIR: Path = Path(__file__).parent

    INPUT_DIR: Path = BASE_DIR / "sample_data"
    OUTPUT_DIR: Path = BASE_DIR / "output"
    LOG_DIR: Path = BASE_DIR / "logs"

    DATE_FORMAT: str = "%Y-%m-%d"
    ENCODING: str = "utf-8"
    MAX_NULL_THRESHOLD: float = 0.5
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/__init__.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/__init__.py"] = textwrap.dedent('''\
    """Excel CSV Automator — core modules package."""

    from modules.cleaner import clean_dataframe
    from modules.combiner import combine_files
    from modules.converter import convert_file
    from modules.reporter import generate_report

    __all__ = ["clean_dataframe", "combine_files", "convert_file", "generate_report"]
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/logger.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/logger.py"] = textwrap.dedent('''\
    """Centralized logging module with colored console output.

    Usage:
        from modules.logger import get_logger
        logger = get_logger(__name__)
        logger.info("Ready.")
    """

    import logging
    from pathlib import Path

    from colorama import Fore, Style, init

    init(autoreset=True)


    class _ColoredFormatter(logging.Formatter):
        """Formatter that adds ANSI color codes based on log level."""

        _COLORS: dict[int, str] = {
            logging.DEBUG: Fore.CYAN,
            logging.INFO: Fore.GREEN,
            logging.WARNING: Fore.YELLOW,
            logging.ERROR: Fore.RED,
            logging.CRITICAL: Fore.MAGENTA,
        }

        def format(self, record: logging.LogRecord) -> str:  # noqa: A003
            color = self._COLORS.get(record.levelno, "")
            record.levelname = f"{color}{record.levelname}{Style.RESET_ALL}"
            record.msg = f"{color}{record.msg}{Style.RESET_ALL}"
            return super().format(record)


    def get_logger(name: str) -> logging.Logger:
        """Return a configured logger with file and colored console handlers.

        Args:
            name: Logger name, typically ``__name__``.

        Returns:
            A fully configured :class:`logging.Logger` instance.
        """
        # Evitar duplicar handlers si el logger ya existe
        logger = logging.getLogger(name)
        if logger.handlers:
            return logger

        logger.setLevel(logging.DEBUG)

        # Importación diferida para evitar dependencia circular
        import config  # noqa: PLC0415

        # Crear directorio de logs si no existe
        config.LOG_DIR.mkdir(parents=True, exist_ok=True)
        log_file: Path = config.LOG_DIR / "app.log"

        # Handler de archivo — nivel DEBUG
        file_handler = logging.FileHandler(log_file, encoding=config.ENCODING)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(
            logging.Formatter(
                fmt="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
                datefmt=config.DATE_FORMAT,
            )
        )

        # Handler de consola — nivel INFO con colores
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(
            _ColoredFormatter(
                fmt="%(asctime)s | %(levelname)-8s | %(message)s",
                datefmt="%H:%M:%S",
            )
        )

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
        return logger
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/cleaner.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/cleaner.py"] = textwrap.dedent('''\
    """Data cleaning utilities for CSV and Excel DataFrames.

    Provides ``clean_dataframe`` which applies a standard cleaning pipeline
    and returns a summary of all changes performed.
    """

    from typing import Any

    import pandas as pd

    from modules.logger import get_logger

    logger = get_logger(__name__)


    def clean_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
        """Clean a DataFrame and return it alongside a change summary.

        Cleaning steps:
        - Remove duplicate rows.
        - Remove fully-empty rows.
        - Strip leading/trailing whitespace from string columns.
        - Normalize column names (lowercase, underscores).
        - Detect columns with more than ``MAX_NULL_THRESHOLD`` null rate.

        Args:
            df: Raw input DataFrame.

        Returns:
            A tuple of (cleaned_df, summary_dict).
        """
        import config  # noqa: PLC0415

        summary: dict[str, Any] = {
            "original_rows": len(df),
            "original_columns": len(df.columns),
            "duplicates_removed": 0,
            "empty_rows_removed": 0,
            "columns_normalized": [],
            "high_null_columns": [],
        }

        # 1. Eliminar duplicados exactos
        before = len(df)
        df = df.drop_duplicates()
        summary["duplicates_removed"] = before - len(df)
        logger.debug(f"Duplicates removed: {summary[\'duplicates_removed\']}")

        # 2. Eliminar filas completamente vacías
        before = len(df)
        df = df.dropna(how="all")
        summary["empty_rows_removed"] = before - len(df)
        logger.debug(f"Empty rows removed: {summary[\'empty_rows_removed\']}")

        # 3. Normalizar nombres de columnas
        original_cols = list(df.columns)
        df.columns = pd.Index(
            [str(col).strip().lower().replace(" ", "_") for col in df.columns]
        )
        summary["columns_normalized"] = [
            f"{old} → {new}"
            for old, new in zip(original_cols, df.columns)
            if str(old) != str(new)
        ]

        # 4. Limpiar espacios en columnas de texto
        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].str.strip()

        # 5. Detectar columnas con alto porcentaje de nulos
        null_pct = df.isnull().mean()
        summary["high_null_columns"] = [
            f"{col} ({pct:.1%})"
            for col, pct in null_pct.items()
            if pct > config.MAX_NULL_THRESHOLD
        ]

        summary["final_rows"] = len(df)
        summary["final_columns"] = len(df.columns)

        logger.info(
            f"Cleaning complete — "
            f"{summary[\'duplicates_removed\']} duplicates removed, "
            f"{summary[\'empty_rows_removed\']} empty rows dropped."
        )
        return df, summary
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/combiner.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/combiner.py"] = textwrap.dedent('''\
    """File combining utilities.

    Merges multiple CSV and XLSX files vertically into a single output file,
    adding a ``source_file`` column to track provenance.
    """

    from pathlib import Path

    import pandas as pd

    import config
    from modules.logger import get_logger

    logger = get_logger(__name__)

    _SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({".csv", ".xlsx", ".xls"})


    def combine_files(
        file_paths: list[str | Path],
        output_path: Path | None = None,
    ) -> Path:
        """Combine multiple CSV/XLSX files into one merged CSV.

        Args:
            file_paths: Paths to input files (.csv or .xlsx).
            output_path: Destination path for the merged file.  Defaults to
                ``OUTPUT_DIR/combined_output.csv``.

        Returns:
            Path to the saved combined file.

        Raises:
            FileNotFoundError: If any input file does not exist.
            ValueError: If a file format is unsupported or all files are empty.
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / "combined_output.csv"

        dataframes: list[pd.DataFrame] = []

        for raw_path in file_paths:
            path = Path(raw_path)

            if not path.exists():
                raise FileNotFoundError(f"File not found: {path}")

            ext = path.suffix.lower()
            if ext not in _SUPPORTED_EXTENSIONS:
                raise ValueError(
                    f"Unsupported format \'{ext}\'. Supported: "
                    + ", ".join(sorted(_SUPPORTED_EXTENSIONS))
                )

            # Cargar según extensión
            if ext == ".csv":
                df = pd.read_csv(path, encoding=config.ENCODING)
            else:
                df = pd.read_excel(path)

            if df.empty:
                logger.warning(f"Skipping empty file: {path.name}")
                continue

            df["source_file"] = path.name
            dataframes.append(df)
            logger.info(f"Loaded {path.name} — {len(df):,} rows")

        if not dataframes:
            raise ValueError("No valid data found in the provided files.")

        combined = pd.concat(dataframes, ignore_index=True, join="outer")

        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        combined.to_csv(output_path, index=False, encoding=config.ENCODING)

        logger.info(
            f"Combined {len(dataframes)} file(s) → {output_path.name} "
            f"({len(combined):,} total rows)"
        )
        return output_path
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/converter.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/converter.py"] = textwrap.dedent('''\
    """File format conversion utilities.

    Converts between CSV, Excel (XLSX), and JSON, preserving the original
    filename stem and saving output to ``OUTPUT_DIR``.
    """

    from pathlib import Path

    import pandas as pd

    import config
    from modules.logger import get_logger

    logger = get_logger(__name__)

    SUPPORTED_FORMATS: frozenset[str] = frozenset({"csv", "excel", "json"})
    _READABLE_EXTENSIONS: frozenset[str] = frozenset({".csv", ".xlsx", ".xls", ".json"})


    def convert_file(input_path: str | Path, output_format: str) -> Path:
        """Convert a file to the specified format.

        Args:
            input_path: Path to the source file.
            output_format: Target format — ``"csv"``, ``"excel"``, or ``"json"``.

        Returns:
            Path to the converted output file.

        Raises:
            FileNotFoundError: If ``input_path`` does not exist.
            ValueError: If ``output_format`` is unsupported or the file cannot be read.
        """
        input_path = Path(input_path)
        output_format = output_format.lower().strip()

        if not input_path.exists():
            raise FileNotFoundError(f"File not found: {input_path}")

        if output_format not in SUPPORTED_FORMATS:
            raise ValueError(
                f"Unsupported output format \'{output_format}\'. "
                f"Choose from: {', '.join(sorted(SUPPORTED_FORMATS))}"
            )

        ext = input_path.suffix.lower()
        if ext not in _READABLE_EXTENSIONS:
            raise ValueError(f"Cannot read format \'{ext}\'.")

        # Leer archivo de entrada
        if ext == ".csv":
            df = pd.read_csv(input_path, encoding=config.ENCODING)
        elif ext in {".xlsx", ".xls"}:
            df = pd.read_excel(input_path)
        else:
            df = pd.read_json(input_path)

        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        stem = input_path.stem

        # Escribir en el formato destino
        if output_format == "csv":
            output_path = config.OUTPUT_DIR / f"{stem}.csv"
            df.to_csv(output_path, index=False, encoding=config.ENCODING)

        elif output_format == "excel":
            output_path = config.OUTPUT_DIR / f"{stem}.xlsx"
            df.to_excel(output_path, index=False, engine="openpyxl")

        else:  # json
            output_path = config.OUTPUT_DIR / f"{stem}.json"
            df.to_json(output_path, orient="records", indent=2, force_ascii=False)

        logger.info(f"Converted: {input_path.name} → {output_path.name}")
        return output_path
''')

# ──────────────────────────────────────────────────────────────────────────────
# modules/reporter.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["modules/reporter.py"] = textwrap.dedent('''\
    """Professional Excel report generator.

    Creates a multi-sheet XLSX report with:
    - **Data** — raw dataset with styled headers.
    - **Summary** — row/column counts, null rates, descriptive statistics.
    - **Null Analysis** — per-column null breakdown with status flags.
    """

    from pathlib import Path

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    import config
    from modules.logger import get_logger

    logger = get_logger(__name__)

    # Paleta de colores corporativos
    _DARK_BLUE = "1F4E79"
    _MID_BLUE = "2E75B6"
    _LIGHT_BLUE = "D6E4F0"
    _WHITE = "FFFFFF"


    # ─── Helpers de formato ──────────────────────────────────────────────────────

    def _header_cell(ws: Worksheet, row: int, col: int, value: object, bg: str = _DARK_BLUE) -> None:
        """Write a bold white-on-dark-blue header cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color=_WHITE, size=11)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    def _auto_fit(ws: Worksheet) -> None:
        """Adjust column widths based on content length."""
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 45)


    def _title(ws: Worksheet, text: str, row: int = 1) -> None:
        """Write a large section title."""
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=14, color=_DARK_BLUE)


    # ─── Sheets ─────────────────────────────────────────────────────────────────

    def _write_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
        """Sheet 1 — raw data with styled headers and frozen top row."""
        ws: Worksheet = wb.create_sheet("Data")

        for col_idx, header in enumerate(df.columns, 1):
            _header_cell(ws, 1, col_idx, header)

        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Zebra striping en filas pares
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid"
                    )

        ws.freeze_panes = "A2"
        _auto_fit(ws)


    def _write_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
        """Sheet 2 — dataset statistics and descriptive numeric summary."""
        ws: Worksheet = wb.create_sheet("Summary")
        _title(ws, "Dataset Overview", row=1)

        stats: list[tuple[str, object]] = [
            ("Total Rows", len(df)),
            ("Total Columns", len(df.columns)),
            ("Total Cells", df.size),
            ("Total Null Values", int(df.isnull().sum().sum())),
            ("Overall Null Rate", f"{df.isnull().mean().mean():.1%}"),
            ("Duplicate Rows", int(df.duplicated().sum())),
            ("Memory Usage (KB)", round(df.memory_usage(deep=True).sum() / 1024, 2)),
        ]

        _header_cell(ws, 3, 1, "Metric", _MID_BLUE)
        _header_cell(ws, 3, 2, "Value", _MID_BLUE)

        for i, (label, value) in enumerate(stats, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        # Unique-values per column
        start_uniq = len(stats) + 6
        _title(ws, "Unique Values per Column", row=start_uniq)
        _header_cell(ws, start_uniq + 2, 1, "Column", _MID_BLUE)
        _header_cell(ws, start_uniq + 2, 2, "Unique Values", _MID_BLUE)
        for i, col in enumerate(df.columns, start_uniq + 3):
            ws.cell(row=i, column=1, value=col).font = Font(bold=True)
            ws.cell(row=i, column=2, value=int(df[col].nunique()))

        # Estadísticas descriptivas para columnas numéricas
        numeric_df = df.select_dtypes(include="number")
        if not numeric_df.empty:
            desc = numeric_df.describe().reset_index()
            start_desc = start_uniq + len(df.columns) + 6
            _title(ws, "Numeric Column Statistics", row=start_desc)
            for col_idx, col in enumerate(desc.columns, 1):
                _header_cell(ws, start_desc + 2, col_idx, col, _MID_BLUE)
            for row_idx, row in enumerate(desc.itertuples(index=False), start_desc + 3):
                for col_idx, value in enumerate(row, 1):
                    try:
                        ws.cell(row=row_idx, column=col_idx, value=round(float(value), 4))
                    except (TypeError, ValueError):
                        ws.cell(row=row_idx, column=col_idx, value=str(value))

        _auto_fit(ws)


    def _write_null_analysis_sheet(wb: Workbook, df: pd.DataFrame) -> None:
        """Sheet 3 — per-column null analysis with status indicators."""
        ws: Worksheet = wb.create_sheet("Null Analysis")
        _title(ws, "Null Value Analysis", row=1)

        headers = ["Column", "Total Nulls", "Null %", "Non-Null Count", "Unique Values", "Data Type", "Status"]
        for col_idx, header in enumerate(headers, 1):
            _header_cell(ws, 3, col_idx, header)

        for row_idx, col in enumerate(df.columns, 4):
            null_count = int(df[col].isnull().sum())
            null_pct = df[col].isnull().mean()
            non_null = int(df[col].notna().sum())
            unique = int(df[col].nunique())
            dtype = str(df[col].dtype)
            status = "HIGH NULLS" if null_pct > config.MAX_NULL_THRESHOLD else "OK"
            status_color = "C00000" if null_pct > config.MAX_NULL_THRESHOLD else "375623"

            ws.cell(row=row_idx, column=1, value=col).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=null_count)
            ws.cell(row=row_idx, column=3, value=f"{null_pct:.1%}")
            ws.cell(row=row_idx, column=4, value=non_null)
            ws.cell(row=row_idx, column=5, value=unique)
            ws.cell(row=row_idx, column=6, value=dtype)
            status_cell = ws.cell(row=row_idx, column=7, value=status)
            status_cell.font = Font(bold=True, color=status_color)

        _auto_fit(ws)


    # ─── Public API ──────────────────────────────────────────────────────────────

    def generate_report(df: pd.DataFrame, output_path: Path | None = None) -> Path:
        """Generate a professional multi-sheet Excel report.

        Args:
            df: Source DataFrame to report on.
            output_path: Destination ``.xlsx`` path. Defaults to
                ``OUTPUT_DIR/report.xlsx``.

        Returns:
            Path to the generated report file.
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / "report.xlsx"

        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Eliminar la hoja por defecto de openpyxl
        wb.remove(wb.active)  # type: ignore[arg-type]

        _write_data_sheet(wb, df)
        _write_summary_sheet(wb, df)
        _write_null_analysis_sheet(wb, df)

        wb.save(output_path)
        logger.info(f"Report generated: {output_path.name} ({len(df):,} rows, {len(df.columns)} columns)")
        return output_path
''')

# ──────────────────────────────────────────────────────────────────────────────
# main.py
# ──────────────────────────────────────────────────────────────────────────────
FILES["main.py"] = textwrap.dedent('''\
    """Excel CSV Automator — CLI entry point.

    Run ``python main.py --help`` for usage information.
    """

    import argparse
    import sys
    from pathlib import Path

    import pandas as pd
    from colorama import Fore, Style, init

    import config
    from modules.logger import get_logger

    init(autoreset=True)

    logger = get_logger("main")

    _BANNER = (
        f"\\n{Fore.CYAN}"
        "╔══════════════════════════════════════════════════════╗\\n"
        "║      Excel & CSV Automator  ·  v1.0.0                ║\\n"
        "║      Professional Data Automation Toolkit            ║\\n"
        "╚══════════════════════════════════════════════════════╝"
        f"{Style.RESET_ALL}\\n"
    )


    # ─── Output helpers ──────────────────────────────────────────────────────────

    def _ok(msg: str) -> None:
        print(f"\\n{Fore.GREEN}  ✅  {msg}{Style.RESET_ALL}")


    def _err(msg: str) -> None:
        print(f"\\n{Fore.RED}  ❌  {msg}{Style.RESET_ALL}")


    def _info(msg: str) -> None:
        print(f"{Fore.CYAN}  ➜   {msg}{Style.RESET_ALL}")


    def _divider() -> None:
        print(f"{Fore.CYAN}  {'─' * 52}{Style.RESET_ALL}")


    def _load_dataframe(path: Path) -> pd.DataFrame:
        """Read CSV or XLSX into a DataFrame."""
        ext = path.suffix.lower()
        if ext == ".csv":
            return pd.read_csv(path, encoding=config.ENCODING)
        elif ext in {".xlsx", ".xls"}:
            return pd.read_excel(path)
        raise ValueError(f"Unreadable format \'{ext}\'. Use .csv or .xlsx.")


    # ─── Sub-commands ────────────────────────────────────────────────────────────

    def cmd_clean(args: argparse.Namespace) -> None:
        """Clean and normalize a CSV/XLSX file."""
        from modules.cleaner import clean_dataframe

        input_path = Path(args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        _info(f"Loading  →  {input_path.name}")
        df = _load_dataframe(input_path)
        _info(f"Cleaning  {len(df):,} rows × {len(df.columns)} columns …")

        cleaned_df, summary = clean_dataframe(df)

        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = config.OUTPUT_DIR / f"cleaned_{input_path.name}"
        if input_path.suffix.lower() == ".csv":
            cleaned_df.to_csv(output_path, index=False, encoding=config.ENCODING)
        else:
            cleaned_df.to_excel(output_path, index=False, engine="openpyxl")

        _ok(f"Saved  →  {output_path}")
        print()
        print(f"{Fore.YELLOW}  📋  Cleaning Summary{Style.RESET_ALL}")
        print(f"     Rows before          {summary[\'original_rows\']:>6}")
        print(f"     Rows after           {summary[\'final_rows\']:>6}")
        print(f"     Duplicates removed   {summary[\'duplicates_removed\']:>6}")
        print(f"     Empty rows dropped   {summary[\'empty_rows_removed\']:>6}")
        if summary["columns_normalized"]:
            print(f"     Columns normalized   {len(summary[\'columns_normalized\']):>6}")
        if summary["high_null_columns"]:
            print(f"\\n  {Fore.YELLOW}⚠   High-null columns:{Style.RESET_ALL}")
            for col in summary["high_null_columns"]:
                print(f"      • {col}")


    def cmd_combine(args: argparse.Namespace) -> None:
        """Combine multiple CSV/XLSX files into one."""
        from modules.combiner import combine_files

        file_paths = [Path(p) for p in args.inputs]
        _info(f"Combining {len(file_paths)} file(s) …")

        output_path = config.OUTPUT_DIR / "combined_output.csv"
        result = combine_files(file_paths, output_path)
        _ok(f"Combined file saved  →  {result}")


    def cmd_convert(args: argparse.Namespace) -> None:
        """Convert a file to another format."""
        from modules.converter import convert_file

        input_path = Path(args.input)
        _info(f"Converting  {input_path.name}  →  {args.format.upper()} …")

        result = convert_file(input_path, args.format)
        _ok(f"Converted file saved  →  {result}")


    def cmd_report(args: argparse.Namespace) -> None:
        """Generate a professional Excel report."""
        from modules.reporter import generate_report

        input_path = Path(args.input)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        _info(f"Loading  →  {input_path.name}")
        df = _load_dataframe(input_path)
        _info(f"Generating report for  {len(df):,} rows × {len(df.columns)} columns …")

        output_path = config.OUTPUT_DIR / f"report_{input_path.stem}.xlsx"
        result = generate_report(df, output_path)
        _ok(f"Report saved  →  {result}")
        print(f"\\n  Open {Fore.CYAN}{result}{Style.RESET_ALL} in Excel to explore the results.")


    # ─── CLI setup ───────────────────────────────────────────────────────────────

    def _build_parser() -> argparse.ArgumentParser:
        parser = argparse.ArgumentParser(
            prog="excel_csv_automator",
            description="Professional Excel & CSV Automation Toolkit",
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog=(
                "Examples:\\n"
                "  python main.py clean   --input sample_data/sample_input.csv\\n"
                "  python main.py combine --inputs file1.csv file2.xlsx\\n"
                "  python main.py convert --input sample_data/sample_input.csv --format json\\n"
                "  python main.py report  --input sample_data/sample_input.csv\\n"
            ),
        )
        sub = parser.add_subparsers(dest="command", required=True, title="commands")

        # clean
        p_clean = sub.add_parser("clean", help="Remove duplicates, fix nulls, normalize columns")
        p_clean.add_argument("--input", required=True, metavar="FILE",
                             help="Path to input .csv or .xlsx file")

        # combine
        p_combine = sub.add_parser("combine", help="Merge multiple CSV/XLSX files into one")
        p_combine.add_argument("--inputs", nargs="+", required=True, metavar="FILE",
                               help="Paths to two or more input files")

        # convert
        p_convert = sub.add_parser("convert", help="Convert a file to csv / excel / json")
        p_convert.add_argument("--input", required=True, metavar="FILE",
                               help="Path to input file")
        p_convert.add_argument("--format", required=True, choices=["csv", "excel", "json"],
                               metavar="FORMAT", help="Target format: csv | excel | json")

        # report
        p_report = sub.add_parser("report", help="Generate a professional multi-sheet Excel report")
        p_report.add_argument("--input", required=True, metavar="FILE",
                              help="Path to input .csv or .xlsx file")

        return parser


    def main() -> None:
        """CLI entry point."""
        print(_BANNER)
        parser = _build_parser()
        args = parser.parse_args()

        dispatch = {
            "clean":   cmd_clean,
            "combine": cmd_combine,
            "convert": cmd_convert,
            "report":  cmd_report,
        }

        try:
            dispatch[args.command](args)
        except FileNotFoundError as exc:
            _err(str(exc))
            logger.error(str(exc))
            sys.exit(1)
        except ValueError as exc:
            _err(str(exc))
            logger.error(str(exc))
            sys.exit(1)
        except KeyboardInterrupt:
            print(f"\\n{Fore.YELLOW}  Interrupted by user.{Style.RESET_ALL}\\n")
            sys.exit(0)
        except Exception as exc:
            _err(f"Unexpected error: {exc}")
            logger.exception("Unexpected error")
            sys.exit(1)

        _divider()
        print(f"{Fore.GREEN}  ✨  Operation completed successfully!{Style.RESET_ALL}")
        _divider()
        print()


    if __name__ == "__main__":
        main()
''')

# ──────────────────────────────────────────────────────────────────────────────
# requirements.txt
# ──────────────────────────────────────────────────────────────────────────────
FILES["requirements.txt"] = textwrap.dedent('''\
    pandas>=2.0.0
    openpyxl>=3.1.0
    colorama>=0.4.6
    tqdm>=4.66.0
''')

# ──────────────────────────────────────────────────────────────────────────────
# .gitignore
# ──────────────────────────────────────────────────────────────────────────────
FILES[".gitignore"] = textwrap.dedent('''\
    # Python
    __pycache__/
    *.py[cod]
    *.pyo
    *.pyd
    .Python
    *.egg-info/
    dist/
    build/
    .eggs/

    # Virtual environments
    .venv/
    venv/
    env/
    ENV/

    # Generated output (keep folders, ignore files)
    output/*
    !output/.gitkeep
    logs/*
    !logs/.gitkeep

    # IDE
    .vscode/
    .idea/
    *.swp
    *.swo

    # OS
    .DS_Store
    Thumbs.db
    desktop.ini

    # Test artifacts
    .pytest_cache/
    .coverage
    htmlcov/
''')

# ──────────────────────────────────────────────────────────────────────────────
# LICENSE
# ──────────────────────────────────────────────────────────────────────────────
FILES["LICENSE"] = textwrap.dedent('''\
    Copyright © 2025 Santiago Sánchez

    PORTFOLIO & DEMONSTRATION LICENSE

    This project and its source code are made available for portfolio and
    demonstration purposes only.

    Permission is hereby granted, free of charge, to any person obtaining a
    copy of this software, to use, copy, and modify the software solely for
    personal, educational, or evaluation purposes, subject to the following
    conditions:

    1. PROHIBITED USES
       - Unauthorized commercial redistribution, resale, or reproduction of
         this software or any substantial portion of its source code is strictly
         prohibited without the prior written permission of the copyright holder.
       - This software may not be re-packaged and sold as part of a product or
         service without explicit permission.

    2. ATTRIBUTION
       - Any permitted use must retain this copyright notice and license text.

    3. NO WARRANTY
       - This software is provided "as is", without warranty of any kind,
         express or implied, including but not limited to the warranties of
         merchantability, fitness for a particular purpose, and non-infringement.
       - In no event shall the copyright holder be liable for any claim, damages,
         or other liability arising from the use of this software.

    For licensing inquiries or commercial use permissions, please contact the
    copyright holder directly.
''')

# ──────────────────────────────────────────────────────────────────────────────
# demo_commands.txt
# ──────────────────────────────────────────────────────────────────────────────
FILES["demo_commands.txt"] = textwrap.dedent('''\
    ╔══════════════════════════════════════════════════════════╗
    ║         Excel CSV Automator — Demo Commands              ║
    ╚══════════════════════════════════════════════════════════╝

    ── SETUP ─────────────────────────────────────────────────
    # 1. Create and activate a virtual environment
    python -m venv .venv
    .venv\\Scripts\\activate          # Windows
    source .venv/bin/activate        # macOS / Linux

    # 2. Install dependencies
    pip install -r requirements.txt

    ── CLEAN ─────────────────────────────────────────────────
    # Clean the sample CSV (removes duplicates, fixes nulls, normalizes columns)
    python main.py clean --input sample_data/sample_input.csv

    # Clean a custom XLSX file
    python main.py clean --input path/to/your_file.xlsx

    ── COMBINE ───────────────────────────────────────────────
    # Combine two files (creates output/combined_output.csv)
    python main.py combine --inputs sample_data/sample_input.csv output/cleaned_sample_input.csv

    # Combine three files of mixed formats
    python main.py combine --inputs file1.csv file2.xlsx file3.csv

    ── CONVERT ───────────────────────────────────────────────
    # CSV → JSON
    python main.py convert --input sample_data/sample_input.csv --format json

    # CSV → Excel
    python main.py convert --input sample_data/sample_input.csv --format excel

    # Excel → CSV
    python main.py convert --input output/cleaned_sample_input.xlsx --format csv

    ── REPORT ────────────────────────────────────────────────
    # Generate a professional Excel report with 3 analysis sheets
    python main.py report --input sample_data/sample_input.csv

    # Report from cleaned output
    python main.py report --input output/cleaned_sample_input.csv

    ── HELP ──────────────────────────────────────────────────
    python main.py --help
    python main.py clean   --help
    python main.py combine --help
    python main.py convert --help
    python main.py report  --help

    ── FULL DEMO WORKFLOW ────────────────────────────────────
    # Step 1 — Clean raw data
    python main.py clean --input sample_data/sample_input.csv

    # Step 2 — Convert cleaned data to JSON
    python main.py convert --input output/cleaned_sample_input.csv --format json

    # Step 3 — Generate a full analysis report
    python main.py report --input output/cleaned_sample_input.csv
''')

# ──────────────────────────────────────────────────────────────────────────────
# sample_data/sample_input.csv
# ──────────────────────────────────────────────────────────────────────────────
FILES["sample_data/sample_input.csv"] = textwrap.dedent('''\
    id,name,email,age,department,salary,join_date,status
    1,  John Smith  ,john.smith@company.com,28,Engineering,75000,2021-03-15,Active
    2,Maria Garcia,maria.garcia@company.com,34,Marketing,62000,2020-07-22,active
    3,DAVID LEE,david.lee@company.com,45,Engineering,95000,2019-01-10,Active
    4,  Sarah Johnson  ,sarah.j@company.com,29,HR,58000,2022-05-30,ACTIVE
    5,Carlos Mendez,carlos.mendez@company.com,38,Sales,71000,2021-11-08,Active
    6,Emily Chen,emily.chen@company.com,26,,55000,2023-02-14,Active
    7,Robert Taylor,r.taylor@company.com,52,Engineering,110000,2017-09-20,Inactive
    8,  Ana Flores  ,ana.flores@company.com,31,Marketing,67000,2020-12-01,Active
    9,Michael Brown,m.brown@company.com,41,Finance,88000,2018-06-15,Active
    10,Lisa Wang,lisa.wang@company.com,,HR,60000,2022-08-10,Active
    11,  John Smith  ,john.smith@company.com,28,Engineering,75000,2021-03-15,Active
    12,James Wilson,j.wilson@company.com,36,Sales,73000,2020-03-25,active
    13,Patricia Davis,patricia.d@company.com,48,Finance,92000,2016-11-30,Active
    14,Kevin Martinez,kevin.m@company.com,27,Engineering,69000,2023-04-12,Active
    15,Nancy Thompson,nancy.t@company.com,39,,64000,,Active
    16,  Carlos Mendez  ,carlos.mendez@company.com,38,Sales,71000,2021-11-08,Active
    17,Sandra White,sandra.w@company.com,44,HR,61000,2019-07-08,INACTIVE
    18,Daniel Harris,d.harris@company.com,33,Engineering,82000,2020-10-17,Active
    19,Jennifer Clark,jennifer.c@company.com,,,55000,2023-01-25,
    20,Thomas Lewis,thomas.l@company.com,50,Finance,98000,2015-04-30,Active
    21,Rachel Young,rachel.y@company.com,29,Marketing,63000,,Active
    22,  Emily Chen  ,emily.chen@company.com,26,,55000,2023-02-14,Active
''')

# ──────────────────────────────────────────────────────────────────────────────
# screenshots/demo_placeholder.txt
# ──────────────────────────────────────────────────────────────────────────────
FILES["screenshots/demo_placeholder.txt"] = textwrap.dedent('''\
    Place your demo screenshots here.

    Suggested screenshots to capture:
    ─────────────────────────────────
    1. clean_command.png     — Terminal output after running `python main.py clean ...`
    2. report_sheets.png     — Excel report showing all 3 sheets (Data / Summary / Null Analysis)
    3. convert_json.png      — JSON output snippet after format conversion
    4. combine_output.png    — Combined CSV file with source_file column visible

    Tips:
    - Use a dark terminal theme for cleaner-looking screenshots
    - Crop tightly to the relevant area
    - Use 1280×800 or higher resolution
    - Include these in the README Screenshots section
''')

# ──────────────────────────────────────────────────────────────────────────────
# logs/.gitkeep  and  output/.gitkeep
# ──────────────────────────────────────────────────────────────────────────────
FILES["logs/.gitkeep"] = ""
FILES["output/.gitkeep"] = ""

# ──────────────────────────────────────────────────────────────────────────────
# README.md
# ──────────────────────────────────────────────────────────────────────────────
FILES["README.md"] = textwrap.dedent('''\
    <div align="center">

    # 📊 Excel & CSV Automator

    ### Professional Data Automation Toolkit

    ![Python](https://img.shields.io/badge/Python-3.11%2B-blue?logo=python&logoColor=white)
    ![License](https://img.shields.io/badge/License-Portfolio-orange)
    ![Status](https://img.shields.io/badge/Status-Active-brightgreen)
    ![Pandas](https://img.shields.io/badge/Pandas-2.0%2B-150458?logo=pandas)
    ![OpenPyXL](https://img.shields.io/badge/OpenPyXL-3.1%2B-217346)

    > **Stop doing repetitive spreadsheet work manually.**  
    > This tool automates your entire CSV/Excel data pipeline — clean, combine, convert, and report
    > in a single command.

    </div>

    ---

    ## ✨ Features

    | Feature | Description |
    |---|---|
    | 🧹 **Smart Cleaning** | Remove duplicates, fix nulls, normalize headers automatically |
    | 🔗 **File Combiner** | Merge multiple CSV/XLSX files with automatic schema alignment |
    | 🔄 **Format Converter** | Convert between CSV, Excel, and JSON in one command |
    | 📈 **Excel Reporter** | Generate professional 3-sheet analysis reports with formatting |
    | 🎨 **Beautiful CLI** | Color-coded terminal output with clear summaries |
    | 📝 **Auto Logging** | Every operation logged to file with timestamp and detail |
    | 🛡️ **Safe & Validated** | All inputs validated with helpful error messages |

    ---

    ## 📁 Project Structure

    ```
    excel_csv_automator/
    ├── main.py                  # CLI entry point
    ├── config.py                # Global configuration
    ├── requirements.txt
    ├── demo_commands.txt        # Ready-to-run demo commands
    ├── sample_data/
    │   └── sample_input.csv     # Realistic demo dataset (22 rows)
    ├── output/                  # Generated files land here
    ├── logs/
    │   └── app.log              # Auto-generated operation log
    └── modules/
        ├── cleaner.py           # Data cleaning pipeline
        ├── combiner.py          # Multi-file merger
        ├── converter.py         # Format conversion
        ├── reporter.py          # Excel report generator
        └── logger.py            # Centralized logging
    ```

    ---

    ## ⚙️ Installation

    **Requirements:** Python 3.11+

    ```bash
    # 1. Clone the repository
    git clone https://github.com/your-username/excel_csv_automator.git
    cd excel_csv_automator

    # 2. Create and activate a virtual environment (recommended)
    python -m venv .venv
    .venv\\Scripts\\activate        # Windows
    source .venv/bin/activate      # macOS / Linux

    # 3. Install dependencies
    pip install -r requirements.txt
    ```

    ---

    ## 🚀 CLI Usage

    ### 🧹 Clean a file

    Removes duplicates, strips whitespace, normalizes column names, and flags high-null columns.

    ```bash
    python main.py clean --input sample_data/sample_input.csv
    ```

    **Output:**
    ```
    ╔══════════════════════════════════════════════════════╗
    ║      Excel & CSV Automator  ·  v1.0.0                ║
    ╚══════════════════════════════════════════════════════╝

      ➜   Loading  →  sample_input.csv
      ➜   Cleaning  22 rows × 8 columns …

      ✅  Saved  →  output/cleaned_sample_input.csv

      📋  Cleaning Summary
         Rows before               22
         Rows after                18
         Duplicates removed         3
         Empty rows dropped         0
         Columns normalized         0

      ⚠   High-null columns:
          • department (27.3%)
    ```

    ---

    ### 🔗 Combine multiple files

    ```bash
    python main.py combine --inputs file1.csv file2.xlsx file3.csv
    ```

    Merges files vertically, adds a `source_file` column, handles different schemas.  
    Output: `output/combined_output.csv`

    ---

    ### 🔄 Convert format

    ```bash
    # CSV → JSON
    python main.py convert --input sample_data/sample_input.csv --format json

    # CSV → Excel
    python main.py convert --input sample_data/sample_input.csv --format excel

    # Excel → CSV
    python main.py convert --input output/cleaned_sample_input.xlsx --format csv
    ```

    Supported formats: `csv` · `excel` · `json`

    ---

    ### 📈 Generate Excel report

    ```bash
    python main.py report --input sample_data/sample_input.csv
    ```

    Creates `output/report_sample_input.xlsx` with three sheets:

    | Sheet | Contents |
    |---|---|
    | **Data** | Full dataset with styled headers and zebra striping |
    | **Summary** | Row/column counts, null rates, descriptive statistics |
    | **Null Analysis** | Per-column null breakdown with OK / HIGH NULLS flags |

    ---

    ## 📥 Input / Output Examples

    ### Input CSV (raw)

    | id | name | age | department | salary | status |
    |---|---|---|---|---|---|
    | 1 | `  John Smith  ` | 28 | Engineering | 75000 | Active |
    | 11 | `  John Smith  ` | 28 | Engineering | 75000 | Active |  ← duplicate |
    | 6 | Emily Chen | 26 | *(empty)* | 55000 | Active |

    ### Output CSV (cleaned)

    | id | name | age | department | salary | status |
    |---|---|---|---|---|---|
    | 1 | John Smith | 28 | Engineering | 75000 | Active |
    | 6 | Emily Chen | 26 | *(empty)* | 55000 | Active |

    ---

    ## 🛡️ Error Handling

    The tool validates inputs and provides clear feedback:

    ```bash
    # File not found
    ❌  Input file not found: data/missing.csv

    # Unsupported format
    ❌  Unsupported output format \'pdf\'. Choose from: csv, excel, json

    # Empty dataset after combining
    ❌  No valid data found in the provided files.
    ```

    All errors are also written to `logs/app.log` with full stack traces.

    ---

    ## 🎬 Demo Workflow

    Run the full pipeline from raw data to polished report:

    ```bash
    # Step 1 — Clean raw sample data
    python main.py clean --input sample_data/sample_input.csv

    # Step 2 — Convert cleaned data to JSON (API-ready format)
    python main.py convert --input output/cleaned_sample_input.csv --format json

    # Step 3 — Generate a full Excel analysis report
    python main.py report --input output/cleaned_sample_input.csv
    ```

    See `demo_commands.txt` for the full command reference.

    ---

    ## 📸 Screenshots

    > *Add screenshots here after running the demo commands.*

    | Clean Command | Excel Report |
    |---|---|
    | `screenshots/clean_command.png` | `screenshots/report_sheets.png` |

    See `screenshots/demo_placeholder.txt` for guidance on what to capture.

    ---

    ## 🔮 Future Improvements

    - [ ] Interactive TUI mode with `rich` or `textual`
    - [ ] Email report delivery via SMTP
    - [ ] Scheduled automation via `schedule` or cron
    - [ ] Support for Google Sheets API integration
    - [ ] Web dashboard with `streamlit`
    - [ ] Config file support (`.toml` / `.yaml`)
    - [ ] Data profiling with `ydata-profiling`

    ---

    ## 📄 License

    Copyright © 2025 Santiago Sánchez.  
    This project is for portfolio and demonstration purposes.  
    See [LICENSE](LICENSE) for full terms.

    ---

    <div align="center">

    **Built with Python · pandas · openpyxl · colorama**

    *If this project helped you, consider leaving a ⭐ on GitHub!*

    </div>
''')


# ──────────────────────────────────────────────────────────────────────────────
# Bootstrap runner
# ──────────────────────────────────────────────────────────────────────────────
def run() -> None:
    created = 0
    skipped = 0
    errors = 0

    print("\n  🚀  Excel CSV Automator — Project Setup\n")

    for relative_path, content in FILES.items():
        target = BASE / relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists():
            print(f"  ⚡  Skip (exists)  →  {relative_path}")
            skipped += 1
        else:
            try:
                target.write_text(content, encoding="utf-8")
                print(f"  ✅  Created       →  {relative_path}")
                created += 1
            except Exception as exc:
                print(f"  ❌  Error         →  {relative_path}: {exc}")
                errors += 1

    print(f"\n  ─────────────────────────────────────────────")
    print(f"  Created: {created}   Skipped: {skipped}   Errors: {errors}")
    print(f"  Project ready at: {BASE}\n")
    print("  Next steps:")
    print("    cd excel_csv_automator")
    print("    pip install -r requirements.txt")
    print("    python main.py clean --input sample_data/sample_input.csv")
    print("    python main.py report --input sample_data/sample_input.csv\n")


if __name__ == "__main__":
    run()

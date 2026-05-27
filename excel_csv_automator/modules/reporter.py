"""Professional Excel report generator.

Creates a multi-sheet XLSX report with:
- **Data** — raw dataset with styled headers.
- **Summary** — row/column counts, null rates, descriptive statistics.
- **Null Analysis** — per-column null breakdown with status flags.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from modules.logger import get_logger

logger = get_logger(__name__)

# Paleta de colores corporativos
_DARK_BLUE = "1F4E79"
_MID_BLUE = "2E75B6"
_LIGHT_BLUE = "D6E4F0"
_WHITE = "FFFFFF"


# ─── Helpers de formato ──────────────────────────────────────────────────────

def _header_cell(ws: Worksheet, row: int, col: int, value: object, bg: str = _DARK_BLUE) -> None:
    """Write a bold white-on-dark-blue header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_fit(ws: Worksheet) -> None:
    """Adjust column widths based on content length."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 45)


def _title(ws: Worksheet, text: str, row: int = 1) -> None:
    """Write a large section title."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color=_DARK_BLUE)


# ─── Sheets ─────────────────────────────────────────────────────────────────

def _write_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 1 — raw data with styled headers and frozen top row."""
    ws: Worksheet = wb.create_sheet("Data")

    for col_idx, header in enumerate(df.columns, 1):
        _header_cell(ws, 1, col_idx, header)

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Zebra striping en filas pares
            if row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid"
                )

    ws.freeze_panes = "A2"
    _auto_fit(ws)


def _write_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 2 — dataset statistics and descriptive numeric summary."""
    ws: Worksheet = wb.create_sheet("Summary")
    _title(ws, "Dataset Overview", row=1)

    stats: list[tuple[str, object]] = [
        ("Total Rows", len(df)),
        ("Total Columns", len(df.columns)),
        ("Total Cells", df.size),
        ("Total Null Values", int(df.isnull().sum().sum())),
        ("Overall Null Rate", f"{df.isnull().mean().mean():.1%}"),
        ("Duplicate Rows", int(df.duplicated().sum())),
        ("Memory Usage (KB)", round(df.memory_usage(deep=True).sum() / 1024, 2)),
    ]

    _header_cell(ws, 3, 1, "Metric", _MID_BLUE)
    _header_cell(ws, 3, 2, "Value", _MID_BLUE)

    for i, (label, value) in enumerate(stats, 4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Unique-values per column
    start_uniq = len(stats) + 6
    _title(ws, "Unique Values per Column", row=start_uniq)
    _header_cell(ws, start_uniq + 2, 1, "Column", _MID_BLUE)
    _header_cell(ws, start_uniq + 2, 2, "Unique Values", _MID_BLUE)
    for i, col in enumerate(df.columns, start_uniq + 3):
        ws.cell(row=i, column=1, value=col).font = Font(bold=True)
        ws.cell(row=i, column=2, value=int(df[col].nunique()))

    # Estadísticas descriptivas para columnas numéricas
    numeric_df = df.select_dtypes(include="number")
    if not numeric_df.empty:
        desc = numeric_df.describe().reset_index()
        start_desc = start_uniq + len(df.columns) + 6
        _title(ws, "Numeric Column Statistics", row=start_desc)
        for col_idx, col in enumerate(desc.columns, 1):
            _header_cell(ws, start_desc + 2, col_idx, col, _MID_BLUE)
        for row_idx, row in enumerate(desc.itertuples(index=False), start_desc + 3):
            for col_idx, value in enumerate(row, 1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=round(float(value), 4))
                except (TypeError, ValueError):
                    ws.cell(row=row_idx, column=col_idx, value=str(value))

    _auto_fit(ws)


def _write_null_analysis_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 3 — per-column null analysis with status indicators."""
    ws: Worksheet = wb.create_sheet("Null Analysis")
    _title(ws, "Null Value Analysis", row=1)

    headers = ["Column", "Total Nulls", "Null %", "Non-Null Count", "Unique Values", "Data Type", "Status"]
    for col_idx, header in enumerate(headers, 1):
        _header_cell(ws, 3, col_idx, header)

    for row_idx, col in enumerate(df.columns, 4):
        null_count = int(df[col].isnull().sum())
        null_pct = df[col].isnull().mean()
        non_null = int(df[col].notna().sum())
        unique = int(df[col].nunique())
        dtype = str(df[col].dtype)
        status = "HIGH NULLS" if null_pct > config.MAX_NULL_THRESHOLD else "OK"
        status_color = "C00000" if null_pct > config.MAX_NULL_THRESHOLD else "375623"

        ws.cell(row=row_idx, column=1, value=col).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=null_count)
        ws.cell(row=row_idx, column=3, value=f"{null_pct:.1%}")
        ws.cell(row=row_idx, column=4, value=non_null)
        ws.cell(row=row_idx, column=5, value=unique)
        ws.cell(row=row_idx, column=6, value=dtype)
        status_cell = ws.cell(row=row_idx, column=7, value=status)
        status_cell.font = Font(bold=True, color=status_color)

    _auto_fit(ws)


# ─── Public API ──────────────────────────────────────────────────────────────

def generate_report(df: pd.DataFrame, output_path: Path | None = None) -> Path:
    """Generate a professional multi-sheet Excel report.

    Args:
        df: Source DataFrame to report on.
        output_path: Destination ``.xlsx`` path. Defaults to
            ``OUTPUT_DIR/report.xlsx``.

    Returns:
        Path to the generated report file.
    """
    if output_path is None:
        output_path = config.OUTPUT_DIR / "report.xlsx"

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Eliminar la hoja por defecto de openpyxl
    wb.remove(wb.active)  # type: ignore[arg-type]

    _write_data_sheet(wb, df)
    _write_summary_sheet(wb, df)
    _write_null_analysis_sheet(wb, df)

    wb.save(output_path)
    logger.info(f"Report generated: {output_path.name} ({len(df):,} rows, {len(df.columns)} columns)")
    return output_path
